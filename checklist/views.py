from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import CategorySerializer, BranchSerializer, ScoreSerializer
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404        
from django.db import IntegrityError
from .models import Branch, Category, Item, Profile, Score
from .forms import AuditForm
from .models import Audit, AuditDetail
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Font
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.conf import settings
import os
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib.auth import login
from django.contrib.auth.backends import ModelBackend
from .models import Branch
from django.contrib.auth.models import User
from django.utils.crypto import constant_time_compare
from django.utils import timezone
from django.db.models import Prefetch


 

def audit_input_view(request):
    """
    Audit ma'lumotlarini kiritish sahifasi.
    Har bir Item uchun: score_ITEMID va image_ITEMID qabul qiladi.
    """

    items = Item.objects.select_related("category").all()

    if request.method == "POST":
        form = AuditForm(request.POST, request.FILES)

        if form.is_valid():
            branch = form.cleaned_data["branch"]

            # Barcha itemlar bo'yicha aylanib chiqamiz
            for item in items:
                score_value = request.POST.get(f"score_{item.id}")
                image_value = request.FILES.get(f"image_{item.id}")

                if score_value is None:
                    continue  # Agar ball tanlanmagan bo'lsa o‘tkazib yuboramiz

                try:
                    score_obj, created = Score.objects.update_or_create(
                        branch=branch,
                        item=item,
                        defaults={
                            "score": int(score_value),
                            "image": image_value
                        }
                    )
                except IntegrityError:
                    pass

            return redirect("audit_result")

    else:
        form = AuditForm()

    return render(request, "audit_form.html", {
        "form": form,
        "items": items,
        "score_choices": [
            (0, "Нет"),
            (1, "Частично"),
            (2, "Есть")
        ],
        
        
    })


# ----------------- MARKDOWN PAGE -----------------
 
def audit_form_view(request):
    """Formani ko'rsatish va POST so'rovini qabul qilish."""
    
    if request.method == 'POST':
        
        # 0. Umumiy ma'lumotlarni olish
        filial_nomi = request.POST.get('filial_nomi')
        # image_file = request.FILES.get('image_...')
        total_percentage_str = request.POST.get('total_percentage')
        print("--- POST MA'LUMOTLARI ---")
        print("Form Ma'lumotlari (request.POST):", request.POST)
        print("Fayl Ma'lumotlari (request.FILES):", request.FILES)
        print("-------------------------")
        if not filial_nomi:
            return HttpResponse("Ошибка: название филиала не указано", status=400)
        try:
            total_percentage = float(total_percentage_str) if total_percentage_str else None
        except ValueError:
            total_percentage = None
            
        # 1. Asosiy Audit yozuvini yaratish (Tranzaksiya ichida)

        try:
             
            with transaction.atomic():

                auditor = request.session.get("auditor_username")

                new_audit = Audit.objects.create(
                    filial_nomi=filial_nomi,
                    total_percentage=total_percentage,
                    auditor=auditor
                )

                for key in request.POST:
                    if key.startswith('score_'):
                        band_id = key.replace('score_', '')
                        score_str = request.POST.get(key)

                        try:
                            score = int(score_str)
                        except ValueError:
                            continue

                        image_key = f'image_{band_id}_1'
                        image_file = request.FILES.get(image_key)

                        AuditDetail.objects.create(
                            audit=new_audit,
                            band_id=band_id,
                            score=score,
                            image=image_file
                        )

            
            # Agar hamma narsa muvaffaqiyatli saqlansa
            return redirect('audit_success') # 👈 'audit_success' URL nomini ishlatamiz
        
        except Exception as e:
            # Agar DB tranzaksiyasida xatolik yuz bersa
            print(f"Xatolik: {e}")
            # return redirect('/audit/success/') 
            return HttpResponse(f"Произошла ошибка при сохранении данных {e}", status=500)

    # GET so'rovi bo'lsa (formani ko'rsatish)
    return render(request, 'worker.html', {"show_action_bar": True})
  


class CategoryListView(generics.ListAPIView):
    queryset = Category.objects.all().prefetch_related("items")
    serializer_class = CategorySerializer

class BranchListCreateView(generics.ListCreateAPIView):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer

class ScoreCreateUpdateView(APIView):
    """
    Create or update a score for branch+item pair.
    Accepts: branch (id), item (id), score (int), image (file, optional)
    """
    def post(self, request, *args, **kwargs):
        branch_id = request.data.get("branch")
        item_id = request.data.get("item")
        score_val = request.data.get("score")
        image = request.FILES.get("image")

        if not (branch_id and item_id and score_val is not None):
            return Response({"detail":"branch, item, score are required"}, status=status.HTTP_400_BAD_REQUEST)

        branch = get_object_or_404(Branch, pk=branch_id)
        item = get_object_or_404(Item, pk=item_id)

        obj, created = Score.objects.get_or_create(branch=branch, item=item, defaults={"score": score_val})
        obj.score = score_val
        if image:
            obj.image = image
        obj.save()
        return Response(ScoreSerializer(obj).data)




def audit_success_view(request):
    """Ma'lumotlar muvaffaqiyatli saqlanganidan keyingi sahifa."""
    context = {
        'message': "Данные аудита успешно сохранены!"
    }
    return render(request, 'success.html', context)




 
def audit_details_page(request):
    queryset = Audit.objects.all()

    # Filial bo'yicha filter
    filial = request.GET.get("filial")
    if filial and filial != "all":
        queryset = queryset.filter(filial_nomi=filial)

    date_from = request.GET.get("from")
    date_to = request.GET.get("to")

    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    queryset = queryset.order_by("-id")

    # HTML uchun  foiz uzarish
    for audit in queryset:
        audit.percent = audit.total_percentage

 

    all_filials = Audit.objects.values_list("filial_nomi", flat=True).distinct()

    return render(request, "audit_page.html", {
        "audits": queryset,
        "all_filials": all_filials,
        "show_action_bar": True
    })
  
 

def generate_excel(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Аудиты"

    # Header qator
    headers = ["Филиал", "Дата", "Процент (%)", "Аудитор"]
    ws.append(headers)

    header_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center

    # Data qatorlar
    row_num = 2
    for a in qs:
        ws.append([
            a.filial_nomi,
            timezone.localtime(a.created_at).strftime("%d-%m-%Y %H:%M"),
            a.total_percentage,
            a.auditor or "-",   # ✅ faqat DB’dan
        ])

        for col in range(1, 5):
            ws.cell(row=row_num, column=col).alignment = center
        row_num += 1

    # Ustun kengligi
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="audits.xlsx"'
    wb.save(response)
    return response

def export_excel(request):
    """
        Параметры GET:
          - filial: "all" или название филиала
          - from: ГГГГ-ММ-ДД (например 2025-12-01)
          - to:   ГГГГ-ММ-ДД
    """
    qs = Audit.objects.all().order_by("-created_at")

    filial = request.GET.get("filial")
    from_date = request.GET.get("from")
    to_date = request.GET.get("to")

    if filial and filial != "all":
        qs = qs.filter(filial_nomi=filial)

    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)

    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)

    return generate_excel(qs)




def export_pdf(request):
    audits = Audit.objects.all().order_by("-created_at")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="audits.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Auditlar ro‘yxati</b>", styles["Title"]))
    elements.append(Spacer(1, 20))

    table_data = [["Filial", "Sana", "Foiz (%)", "Auditor"]]
    for a in audits:
        table_data.append([
            a.filial_nomi,
            a.created_at.strftime("%d-%m-%Y %H:%M"),
            f"{a.total_percentage}%",
            a.auditor or "-",          # ✅ DB
        ])

    table = Table(table_data, colWidths=[160, 120, 70, 90])
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
    ]))

    elements.append(table)
    doc.build(elements)
    return response



def generate_audit_detail_pdf(audit):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="audit_{audit.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30, 
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    elements = []

    created_local = timezone.localtime(audit.created_at)

    # ✅ HEADER
    header = Paragraph(
        f"""
        <b>Audit Tafsilotlari</b><br/>
        Auditor: <b>{audit.auditor or '-'}</b><br/>
        Filial: <b>{audit.filial_nomi}</b><br/>
        Audit vaqti: {created_local.strftime('%d-%m-%Y %H:%M')}<br/>
        Umumiy foiz: <b>{audit.total_percentage}%</b>
        """,
        styles["Title"]
    )
    elements.append(header)
    elements.append(Spacer(1, 16))

    # ✅ TABLE DATA
    table_data = [["Band nomi", "Ball", "Rasm"]]

    for d in audit.details.all():
        band_text = Paragraph(str(d.band_id), styles["Normal"])

        img = "-"
        if d.image:
            img_path = os.path.join(settings.MEDIA_ROOT, d.image.name)
            if os.path.exists(img_path):
                img = Image(img_path, width=70, height=70)

        table_data.append([band_text, str(d.score), img])

    table = Table(table_data, colWidths=[280, 50, 140])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)

    doc.build(elements)
    return response







def audit_filial_detail(request, filial_nomi):
    audits = Audit.objects.filter(filial_nomi=filial_nomi).order_by("-created_at")

    audit_data = []

    for audit in audits:
        details = audit.details.all()

        bands = []
        for d in details:
            bands.append({
                "band_name": d.band_id,     # string: "hujjatlar"
                "score": d.score,
                "images": [d.image.url] if d.image else []
            })

        audit_data.append({
            "audit": audit,
            "bands": bands,
            "percent": audit.total_percentage  # ✅ TO‘G‘RIDAN-TO‘G‘RI
        })

    return render(request, "audit_filial_detail.html", {
        "filial_name": filial_nomi,
        "audit_data": audit_data,
        "show_action_bar": True
    })


@require_POST
def audit_delete(request, id):
    audit = get_object_or_404(Audit, id=id)
    audit.delete()
    return redirect(request.META.get("HTTP_REFERER", "/"))



 
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def export_audit_detail_excel(request, audit_id):
    audit = get_object_or_404(Audit, id=audit_id)
    qs = audit.details.all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Аудит {audit.id}"

    # ✅ Tekislash uslublarini aniqlab olamiz
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # =================
    #  Header ma'lumotlari (Tepada)
    # =================
    ws.merge_cells('A1:C1')
    ws['A1'] = f"ID аудита: {audit.id}"
    ws['A2'] = f"Филиал: {audit.filial_nomi}"
    ws['A3'] = f"Аудитор: {audit.auditor or '-'}"          
    ws['A4'] = f"Общий процент: {audit.total_percentage}%"   

    for cell_name in ['A1', 'A2', 'A3', 'A4']:
        ws[cell_name].font = Font(bold=True, size=12)
        ws[cell_name].alignment = left_wrap

    # =================
    #  Jadval sarlavhalari (Headers)
    # =================
    headers = ["Название пункта", "Балл", "Изображение"]
    start_row = 6 
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        # Sarlavhalar odatda o'rtada tursa chiroyli ko'rinadi
        cell.alignment = center_wrap
        ws.column_dimensions[get_column_letter(col_num)].width = 30 # Kenglikni biroz oshirdik

    # =================
    #  Jadval ma'lumotlari (Data)
    # =================
    row_num = start_row + 1
    for d in qs:
        # Ma'lumotlarni kataklarga yozamiz
        ws.cell(row=row_num, column=1, value=str(d.band_id))
        ws.cell(row=row_num, column=2, value=d.score)
        ws.cell(row=row_num, column=3, value=d.image.url if d.image else "-")

        # ✅ Ustunlar bo'yicha turli tekislashni qo'llaymiz
        for col in range(1, 4):
            if col == 2:
                # Faqat 2-ustun (Балл) o'rtada
                ws.cell(row=row_num, column=col).alignment = center_wrap
            else:
                # 1-ustun (Название) va 3-ustun (Rasm linki) chapdan
                ws.cell(row=row_num, column=col).alignment = left_wrap
        
        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="audit_{audit.id}.xlsx"'
    wb.save(response)
    return response


def export_audit_detail_pdf(request, audit_id):
    audit = get_object_or_404(Audit, id=audit_id)
    return generate_audit_detail_pdf(audit)



def custom_logout(request):
    logout(request)
    return redirect('https://www.google.com/')   

 


def worker_login(request):
    error = None

    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip().lower()
        pin = (request.POST.get('password') or '').strip()

        if not username or not pin:
            error = "Требуется ввод имени пользователя и пароля!"
        elif not settings.GLOBAL_WORKER_PIN:
            error = "На сервере не настроен GLOBAL_WORKER_PIN!"
        elif constant_time_compare(pin, settings.GLOBAL_WORKER_PIN):
            user, created = User.objects.get_or_create(username=username)

            if created or user.has_usable_password():
                user.set_unusable_password()
                user.save()

            Profile.objects.get_or_create(user=user, defaults={"role": "worker"})

            login(request, user, backend="django.contrib.auth.backends.ModelBackend")

            # ✅ MANA SHU QATOR (MUHIM)
            request.session["auditor_username"] = username

            return redirect('audit_form')
        else:
            error = "Неверный пароль!"

    return render(request, 'worker_login.html', {'error': error})


@login_required
def audit_form(request):
    profile = getattr(request.user, "profile", None)
    if not profile or profile.role != "worker":
        return redirect('/')
    return render(request, 'audit_form.html')




 
def low_scores_view(request):
    # filterlar (ixtiyoriy)
    filial = request.GET.get("filial")          # masalan: "Atlas"
    score = request.GET.get("score")           # "0" yoki "1" yoki "all"
    date_from = request.GET.get("from")        # "2025-12-01"
    date_to = request.GET.get("to")            # "2025-12-17"

    audits_qs = Audit.objects.all().order_by("-created_at")

    if filial and filial != "all":
        audits_qs = audits_qs.filter(filial_nomi=filial)

    if date_from:
        audits_qs = audits_qs.filter(created_at__date__gte=date_from)
    if date_to:
        audits_qs = audits_qs.filter(created_at__date__lte=date_to)

    # Past ball detail-larni oldindan olib kelamiz
    detail_qs = AuditDetail.objects.filter(score__in=[0, 1]).order_by("band_id")

    if score in ("0", "1"):
        detail_qs = detail_qs.filter(score=int(score))

    audits_qs = audits_qs.prefetch_related(
        Prefetch("details", queryset=detail_qs, to_attr="low_details")
    )

    # dropdown uchun filiallar
    all_filials = Audit.objects.values_list("filial_nomi", flat=True).distinct()

    return render(request, "low_scores.html", {
        "audits": audits_qs,
        "all_filials": all_filials,
        "selected_filial": filial or "all",
        "selected_score": score or "all",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "show_action_bar": True
    })
